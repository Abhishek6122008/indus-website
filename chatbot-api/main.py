"""
The Indus Group – Chatbot API
Converted from CLI to FastAPI REST server.

Run:
    pip install -r requirements.txt
    uvicorn main:app --reload --port 8000

Endpoints:
    POST /chat            — main chat endpoint (per-session state via session_id)
    GET  /health          — liveness + context check
    POST /refresh-context — force re-scrape of website
    GET  /contacts        — list all saved contacts (admin)
"""

from __future__ import annotations

import csv
import logging
import os
import random
import re
from contextlib import asynccontextmanager
from datetime import datetime
from functools import lru_cache
from typing import Optional
from uuid import uuid4

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Header
from fastapi.middleware.cors import CORSMiddleware
from mistralai.client import Mistral
from pydantic import BaseModel, Field

# ── env & logging ─────────────────────────────────────────────────────────────
load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

# ── config ────────────────────────────────────────────────────────────────────
MISTRAL_API_KEY: str  = os.environ["MISTRAL_API_KEY"]
MODEL:           str  = os.getenv("MISTRAL_MODEL", "mistral-small-latest")
MAX_CHARS_PAGE:  int  = int(os.getenv("MAX_CHARS_PAGE", "5000"))
ALLOWED_ORIGINS: list = os.getenv("ALLOWED_ORIGINS", "http://localhost:3000").split(",")
ADMIN_TOKEN:     str  = os.getenv("ADMIN_TOKEN", "changeme")   # for /contacts endpoint

CONTACTS_CSV  = os.getenv("CONTACTS_CSV", "indus_contacts.csv")
CONTACTS_XLSX = os.getenv("CONTACTS_XLSX", "indus_contacts.xlsx")

SCRAPE_URLS: list[str] = [
    "https://theindusgroup.com/",
    "https://theindusgroup.com/about.html",
    "https://theindusgroup.com/kc.html",
    "https://theindusgroup.com/resourses.html",
    "https://theindusgroup.com/tools.html",
]

# ── xlsx setup (optional) ─────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    log.warning("openpyxl not installed — contacts will fall back to CSV.")

XLSX_HEADERS = ["Timestamp", "Name", "Phone", "Email", "Company", "Topic of Interest"]
HEADER_COLOR = "1A3C5E"
HEADER_FONT  = "FFFFFF"


def _init_xlsx() -> None:
    if os.path.exists(CONTACTS_XLSX):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(XLSX_HEADERS)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color=HEADER_FONT, name="Arial", size=11)
        cell.fill      = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    for col, width in zip(ws.iter_cols(1, len(XLSX_HEADERS)), [20, 18, 16, 28, 22, 24]):
        ws.column_dimensions[col[0].column_letter].width = width
    wb.save(CONTACTS_XLSX)


def save_contact(name: str, phone: str, email: str, company: str, topic: str) -> str:
    ts  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = [ts, name, phone, email, company, topic]
    if XLSX_AVAILABLE:
        _init_xlsx()
        wb = openpyxl.load_workbook(CONTACTS_XLSX)
        ws = wb.active
        ws.append(row)
        fill_color = "EAF2FB" if ws.max_row % 2 == 0 else "FFFFFF"
        for cell in ws[ws.max_row]:
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")
        wb.save(CONTACTS_XLSX)
        return CONTACTS_XLSX
    else:
        write_header = not os.path.exists(CONTACTS_CSV)
        with open(CONTACTS_CSV, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if write_header:
                w.writerow(XLSX_HEADERS)
            w.writerow(row)
        return CONTACTS_CSV


# ── static data ───────────────────────────────────────────────────────────────
BUSINESS_TERMS: dict[str, str] = {
    "msme":       "MSME stands for **Micro, Small and Medium Enterprises** — businesses classified by investment and turnover size. The Indus Group specialises in empowering MSMEs with consulting, finance guidance, and enterprise solutions. 💼",
    "npa":        "NPA stands for **Non-Performing Asset** — a loan where repayment has stopped for 90+ days. The Indus Group helps businesses manage and resolve NPA issues. 📊",
    "erp":        "ERP stands for **Enterprise Resource Planning** — software that integrates core business processes. We advise companies on ERP adoption and digital tools. 🖥️",
    "crm":        "CRM stands for **Customer Relationship Management** — tools to manage client interactions. We help businesses pick and implement the right CRM systems. 🤝",
    "roi":        "ROI stands for **Return on Investment** — a measure of profitability. The Indus Group helps businesses maximise ROI through strategic consulting. 📈",
    "kpi":        "KPI stands for **Key Performance Indicator** — measurable goals that track business success. We help companies define and monitor their KPIs effectively. 🎯",
    "b2b":        "B2B stands for **Business-to-Business** — transactions between companies. The Indus Group primarily serves B2B clients across sectors. 🏢",
    "b2c":        "B2C stands for **Business-to-Consumer** — selling directly to end customers. We advise on both B2B and B2C strategies. 🛒",
    "logistics":  "Logistics refers to the planning and control of movement and storage of goods. The Indus Group offers end-to-end logistics solutions for enterprises. 🚚",
    "enterprise": "An enterprise is a large organisation or business. The Indus Group provides specialised solutions tailored for enterprise-scale operations. 🏛️",
}

PROFANITY: list[str] = [
    'fuck','shit','asshole','bitch','damn','crap','dick','pussy','bastard','cunt',
    'motherfucker','screw you','bc','bsdk','mdc','bkc','chutiya','madarchod',
    'bhenchod','randi','harami','laude','lodu','gandu','suar','kutta','kamina',
    'nalayak','maa ki chut','bhen ki chut','chut','lund','loda','gand',
    'haramzada','haramzaadi','raand','mc',
]

OFF_TOPIC_KEYWORDS: list[str] = [
    'capital of','president of','prime minister','population of',
    'how to cook','recipe','weather','sports','cricket','football',
    'movie','film','song','music','actor','actress','celebrity',
    'python tutorial','java','javascript','html','css','coding','programming',
    'my relationship','love advice','dating','marriage advice',
    'symptom','medicine','doctor','disease','health advice',
    'stock market','bitcoin','crypto','share price','nifty','sensex',
]

FAREWELL_TRIGGERS: list[str] = [
    'bye','goodbye','tata','see you','alvida','au revoir','adios',
    'ok bye','okay bye','chalo bye','ab jaata hoon','ab jaati hoon',
    'cya','take care','nikal','band karo','band kar','khatam',
    'thats all',"that's all",'done','finish','finished','no more questions',
    'nothing else','all good','all done','ok thanks','okay thanks',
    'shukriya bye','dhanyavaad bye','gracias bye',
]

LANG_GREETINGS: dict[str, list[str]] = {
    "hi":    ["नमस्ते! 🙏 The Indus Group के बारे में कुछ जानना हो तो पूछिए।",
              "हेलो! हम Indus Group के बारे में आपकी मदद कर सकते हैं।"],
    "en":    ["Hey there! 👋 How can I help you with The Indus Group today?",
              "Hello! Ask me anything about our services, team, or solutions."],
    "hi_en": ["Hi! Indus Group ke baare mein kuch jaanna hai? Puchho! 😊",
              "Namaste! Main The Indus Group ke baare mein help kar sakta hoon."],
    "fr":    ["Bonjour! Comment puis-je vous aider avec The Indus Group?"],
    "es":    ["¡Hola! ¿Cómo puedo ayudarte con The Indus Group?"],
}

OFF_TOPIC_REPLIES: list[str] = [
    "I'm focused only on The Indus Group topics! 😊 Ask me about our services, team, or solutions.",
    "That's outside my scope — try asking about our logistics, MSME support, or leadership!",
    "Main sirf The Indus Group ke baare mein baat kar sakta hoon. 🙏 Kuch company related poochho!",
]

DONT_KNOW: list[str] = [
    "Hmm, I don't have that detail. Ask me about our services, team, or solutions!",
    "Not sure about that one. Try asking about MSME support, logistics, or our leadership.",
    "Mujhe yeh pata nahi, lekin Indus Group ke services ke baare mein zaroor bata sakta hoon! 😊",
]

CONTACT_PROMPTS: dict[str, str] = {
    "ask_name":    "Before you go, we'd love to stay in touch! 😊\nMay I have your **name** please?",
    "ask_phone":   "Thanks {name}! What's the best **phone number** to reach you?",
    "ask_email":   "Got it! And your **email address**?",
    "ask_company": "Perfect! Which **company or organisation** are you from? (Type 'skip' to skip)",
    "ask_topic":   "Last one — what **topic** were you most interested in today? (e.g. Logistics, MSME, Enterprise Solutions)",
}

GREETING_WORDS = {'hi','hello','hey','namaste','hola','bonjour','salam','namaskar','pranam','hii','sup'}
BAIL_WORDS     = {'cancel','skip all','no thanks','nahi chahiye'}


# ── scraper & context ─────────────────────────────────────────────────────────
def _scrape_url(url: str) -> str:
    try:
        r = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "meta", "link"]):
            tag.decompose()
        return " ".join(soup.get_text(" ", strip=True).split())
    except Exception as exc:
        log.warning("Failed to scrape %s: %s", url, exc)
        return ""


@lru_cache(maxsize=1)
def load_context() -> str:
    chunks = [t for url in SCRAPE_URLS if (t := _scrape_url(url))]
    if not chunks:
        raise RuntimeError("Could not scrape any company page.")
    words = " ".join(chunks).split()
    ctx   = " ".join(words[:3000])
    log.info("Context loaded: %d words", len(words[:3000]))
    return ctx


# ── language detector ─────────────────────────────────────────────────────────
def detect_lang(text: str) -> str:
    if re.search(r"[\u0900-\u097F]", text):
        return "hi"
    t = text.lower()
    if any(x in t for x in ["hindi me","hindi mein","hindi mai","translate to hindi"]):
        return "hi_en"
    if any(x in t for x in ["en francais","en français","translate to french"]):
        return "fr"
    if any(x in t for x in ["en español","en espanol","translate to spanish"]):
        return "es"
    hinglish = ['kya','hai','kaise','mujhe','hum','aap','batao','bolo','karo','nahi','haan',
                'theek','accha','yaar','bhai','dost','matlab','samjho','puchna','mera',
                'isko','samjhao','wala','wali','kar','raha','chahiye','batana']
    if any(w in t.split() for w in hinglish):
        return "hi_en"
    if any(w in t.split() for w in ['bonjour','merci','oui','non','comment','pourquoi']):
        return "fr"
    if any(w in t.split() for w in ['hola','gracias','como','por','favor','donde']):
        return "es"
    return "en"


# ── Mistral helpers ───────────────────────────────────────────────────────────
_mistral = Mistral(api_key=MISTRAL_API_KEY)


def _call_mistral(prompt: str, max_tokens: int = 150) -> str:
    try:
        resp = _mistral.chat.complete(
            model=MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3,
            max_tokens=max_tokens,
        )
        return resp.choices[0].message.content.strip()
    except Exception as exc:
        log.error("Mistral error: %s", exc)
        return ""


def _translate(text: str, lang: str) -> str:
    instructions = {
        "hi":    "Translate into natural Hindi (Devanagari script). Keep **bold** markers.",
        "hi_en": "Translate into Hinglish (Hindi+English, Roman script). Keep **bold** markers.",
        "fr":    "Translate into French. Keep **bold** markers.",
        "es":    "Translate into Spanish. Keep **bold** markers.",
    }
    instr = instructions.get(lang)
    if not instr:
        return text
    result = _call_mistral(f"{instr}\n\nText:\n{text}\n\nTranslation:", max_tokens=200)
    return result or text


def _ask_mistral(question: str, context: str, lang: str) -> str:
    lang_instr = {
        "hi":    "Respond in Hindi (Devanagari script).",
        "hi_en": "Respond in Hinglish (Hindi+English mix, Roman script).",
        "fr":    "Respond in French.",
        "es":    "Respond in Spanish.",
        "en":    "Respond in English.",
    }.get(lang, "Respond in English.")

    prompt = f"""You are a helpful assistant ONLY for The Indus Group company.

RULES:
- ONLY answer questions about The Indus Group — services, team, leadership, solutions, operations.
- If unrelated, politely say you can only help with Indus Group topics.
- Keep answers short: 1-3 sentences max.
- Sound warm and human, not robotic.
- Do NOT say "based on the text" or "according to the document".
- {lang_instr}

COMPANY INFO:
{context}

USER QUESTION: {question}

Answer:"""
    return _call_mistral(prompt, max_tokens=120) or random.choice(DONT_KNOW)


# ── per-session bot state ─────────────────────────────────────────────────────
class SessionState:
    """All mutable state for one chat session."""
    __slots__ = ("warn_count", "collecting_contact", "contact_stage", "contact_data")

    def __init__(self):
        self.warn_count:         int             = 0
        self.collecting_contact: bool            = False
        self.contact_stage:      Optional[str]   = None
        self.contact_data:       dict[str, str]  = {}


# In-memory session store  {session_id: SessionState}
# For production, swap this for Redis or a DB-backed store.
_sessions: dict[str, SessionState] = {}


def get_session(session_id: str) -> SessionState:
    if session_id not in _sessions:
        _sessions[session_id] = SessionState()
    return _sessions[session_id]


# ── bot logic (stateless functions operating on SessionState) ─────────────────
def _has_profanity(text: str) -> bool:
    t = text.lower()
    return any(w in t for w in PROFANITY)


def _is_farewell(text: str) -> bool:
    t = text.lower().strip()
    return t in FAREWELL_TRIGGERS or any(t.startswith(f) for f in FAREWELL_TRIGGERS)


def _is_off_topic(text: str) -> bool:
    t = text.lower()
    return any(kw in t for kw in OFF_TOPIC_KEYWORDS)


def _check_business_term(text: str, lang: str) -> Optional[str]:
    t = text.lower()
    for term, explanation in BUSINESS_TERMS.items():
        if re.search(rf"\b{re.escape(term)}\b", t):
            return _translate(explanation, lang) if lang != "en" else explanation
    return None


def _start_contact_collection(state: SessionState) -> str:
    state.collecting_contact = True
    state.contact_stage      = "ask_name"
    state.contact_data       = {}
    return CONTACT_PROMPTS["ask_name"]


def _handle_contact_step(state: SessionState, user_input: str) -> str:
    val  = user_input.strip()
    skip = val.lower() in {"skip", "na", "n/a", "-", "nahi", "nope", "no"}

    if state.contact_stage == "ask_name":
        if not val or len(val) < 2:
            return "Could you share your name? (Type 'skip' to skip)"
        state.contact_data["name"] = val
        state.contact_stage = "ask_phone"
        return CONTACT_PROMPTS["ask_phone"].format(name=val.split()[0])

    if state.contact_stage == "ask_phone":
        cleaned = re.sub(r"[\s\-\(\)]", "", val)
        if not skip and not re.match(r"^\+?\d{7,15}$", cleaned):
            return "That doesn't look like a valid phone number. Please try again (or type 'skip')."
        state.contact_data["phone"] = val if not skip else "—"
        state.contact_stage = "ask_email"
        return CONTACT_PROMPTS["ask_email"]

    if state.contact_stage == "ask_email":
        if not skip and not re.match(r"^[\w\.\+\-]+@[\w\-]+\.[a-zA-Z]{2,}$", val):
            return "Hmm, that email doesn't look right. Please try again (or type 'skip')."
        state.contact_data["email"] = val if not skip else "—"
        state.contact_stage = "ask_company"
        return CONTACT_PROMPTS["ask_company"]

    if state.contact_stage == "ask_company":
        state.contact_data["company"] = val if not skip else "—"
        state.contact_stage = "ask_topic"
        return CONTACT_PROMPTS["ask_topic"]

    if state.contact_stage == "ask_topic":
        state.contact_data["topic"] = val if not skip else "General Enquiry"
        return _finish_contact(state)

    return "Something went wrong. Let's start over — what's your name?"


def _finish_contact(state: SessionState) -> str:
    d        = state.contact_data
    filepath = save_contact(
        name    = d.get("name",    "—"),
        phone   = d.get("phone",   "—"),
        email   = d.get("email",   "—"),
        company = d.get("company", "—"),
        topic   = d.get("topic",   "—"),
    )
    state.collecting_contact = False
    state.contact_stage      = None
    state.contact_data       = {}
    ext = "spreadsheet" if filepath.endswith(".xlsx") else "CSV"
    return (
        f"✅ Thanks {d.get('name', 'there')}! Your details have been saved.\n"
        f"One of our experts will reach out to you soon. 🙏\n"
        f"Have a great day! 👋"
    )


def bot_respond(state: SessionState, user_input: str, context: str) -> tuple[str, bool]:
    """
    Returns (reply, terminate).
    terminate=True means the session should be ended (3× profanity).
    """
    text = user_input.strip()

    # 1. Profanity guard
    if _has_profanity(text):
        state.warn_count += 1
        if state.warn_count == 1:
            return "Hey, let's keep it respectful please. 🙏 Ask me about The Indus Group.", False
        if state.warn_count == 2:
            return "Last warning — one more and I'll end this chat.", False
        return "I've warned you twice. Ending this session now. Take care.", True

    # 2. Active contact collection
    if state.collecting_contact:
        if text.lower() in BAIL_WORDS:
            state.collecting_contact = False
            state.contact_stage      = None
            state.contact_data       = {}
            return "No problem! Feel free to come back anytime. 👋", False
        return _handle_contact_step(state, text), False

    lang    = detect_lang(text)
    t_lower = text.lower()

    # 3. Greeting
    if t_lower in GREETING_WORDS:
        return random.choice(LANG_GREETINGS.get(lang, LANG_GREETINGS["en"])), False

    # 4. Farewell → start contact collection
    if _is_farewell(text):
        return _start_contact_collection(state), False

    # 5. Business term glossary
    term_answer = _check_business_term(text, lang)
    if term_answer:
        return term_answer, False

    # 6. Off-topic block
    if _is_off_topic(text):
        return random.choice(OFF_TOPIC_REPLIES), False

    # 7. LLM
    return _ask_mistral(text, context, lang), False


# ── lifespan ──────────────────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    log.info("Warming up context cache…")
    try:
        load_context()
        log.info("Context ready.")
    except Exception as exc:
        log.error("Startup scrape failed: %s", exc)
    yield


# ── app ───────────────────────────────────────────────────────────────────────
app = FastAPI(
    title="Indus Group Chatbot API",
    version="2.0.0",
    description="Full-featured RAG chatbot: multilingual, contact capture, profanity guard.",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── schemas ───────────────────────────────────────────────────────────────────
class ChatRequest(BaseModel):
    message:    str           = Field(..., min_length=1, max_length=500)
    session_id: Optional[str] = Field(None, description="Reuse an existing session. Omit to start fresh.")


class ChatResponse(BaseModel):
    reply:      str
    session_id: str
    terminated: bool = False   # True when profanity threshold hit


class HealthResponse(BaseModel):
    status:        str
    context_words: int
    model:         str
    sessions_open: int


class ContactRow(BaseModel):
    timestamp: str
    name:      str
    phone:     str
    email:     str
    company:   str
    topic:     str


# ── routes ────────────────────────────────────────────────────────────────────
@app.get("/health", response_model=HealthResponse, tags=["meta"])
def health():
    try:
        ctx = load_context()
        return HealthResponse(
            status        = "ok",
            context_words = len(ctx.split()),
            model         = MODEL,
            sessions_open = len(_sessions),
        )
    except Exception as exc:
        raise HTTPException(status_code=503, detail=str(exc))


@app.post("/chat", response_model=ChatResponse, tags=["chat"])
def chat(req: ChatRequest):
    """
    Send a message. Pass `session_id` from a previous response to continue
    a conversation (preserves contact-capture state, warn count, etc.).
    Omit it to start a new session.
    """
    try:
        context = load_context()
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc))

    session_id = req.session_id or str(uuid4())
    state      = get_session(session_id)

    reply, terminate = bot_respond(state, req.message, context)

    if terminate:
        # Clean up terminated session
        _sessions.pop(session_id, None)

    return ChatResponse(reply=reply, session_id=session_id, terminated=terminate)


@app.post("/refresh-context", tags=["meta"])
def refresh_context():
    """Force a re-scrape of all pages (clears the in-memory cache)."""
    load_context.cache_clear()
    try:
        ctx = load_context()
        return {"status": "refreshed", "context_words": len(ctx.split())}
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc))


@app.get("/contacts", response_model=list[ContactRow], tags=["admin"])
def list_contacts(x_admin_token: str = Header(...)):
    """
    Return all captured leads. Requires X-Admin-Token header.
    Set ADMIN_TOKEN in your .env to secure this endpoint.
    """
    if x_admin_token != ADMIN_TOKEN:
        raise HTTPException(status_code=403, detail="Invalid admin token.")

    rows: list[ContactRow] = []

    if XLSX_AVAILABLE and os.path.exists(CONTACTS_XLSX):
        wb = openpyxl.load_workbook(CONTACTS_XLSX)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                rows.append(ContactRow(
                    timestamp = str(row[0] or ""),
                    name      = str(row[1] or ""),
                    phone     = str(row[2] or ""),
                    email     = str(row[3] or ""),
                    company   = str(row[4] or ""),
                    topic     = str(row[5] or ""),
                ))
        return rows

    if os.path.exists(CONTACTS_CSV):
        with open(CONTACTS_CSV, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for r in reader:
                rows.append(ContactRow(
                    timestamp = r.get("Timestamp", ""),
                    name      = r.get("Name", ""),
                    phone     = r.get("Phone", ""),
                    email     = r.get("Email", ""),
                    company   = r.get("Company", ""),
                    topic     = r.get("Topic of Interest", ""),
                ))
        return rows

    return []